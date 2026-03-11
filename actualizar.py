"""
actualizar.py  ·  Portfolio Dashboard Updater
=============================================
Lee el Excel de monitoreo, recalcula posiciones FIFO, obtiene precios
frescos via Yahoo Finance, regenera el HTML y lo sube a GitHub Pages.

Uso:
    python actualizar.py                  # corre una vez
    python actualizar.py --watch          # vigila cambios en el Excel
    python actualizar.py --no-github      # solo regenera HTML, no sube
"""

import sys, os, re, json, time, hashlib, subprocess, argparse
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

# ── Dependencias opcionales ────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("ERROR: Falta openpyxl. Ejecuta:  pip install openpyxl")
    sys.exit(1)

try:
    import yfinance as yf
    YFINANCE_OK = True
except ImportError:
    print("AVISO: yfinance no instalado. Precios se tomarán del Excel.")
    print("       Para precios en tiempo real:  pip install yfinance")
    YFINANCE_OK = False

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN  ── edita estas rutas según tu entorno
# ══════════════════════════════════════════════════════════════════════════════

EXCEL_PATH = Path(
    r"C:\Users\vmendoza\OneDrive - DATA SYSTEM & GLOBAL SERVICES S.A.C"
    r"\Escritorio\Victor Mendoza\Portafolio corporativo\RV\RV-Monitoreo_N__3_.xlsx"
)

# Carpeta donde vive este script y el template HTML
BASE_DIR   = Path(__file__).parent
TEMPLATE   = BASE_DIR / "template.html"        # HTML base (sin datos)
OUTPUT     = BASE_DIR / "index.html"           # HTML final para GitHub Pages

# Repositorio GitHub (usuario/repo)  ← cámbialo por el tuyo
GITHUB_REPO = "vmendoza/portfolio-dashboard"   # EDITAR

# Constantes financieras
PEN_RATE   = 3.72
RF_ANNUAL  = 0.04375
RF_MONTHLY = RF_ANNUAL / 12

# Tickers PEN (cotizan en BVL, precios en soles)
PEN_TICKERS = {"MINSURI1", "FERREYC1", "SIDERC1", "ORYGENC1",
               "CPACASC1", "INRETC1", "JDOC", "JAVA", "VOLCABC1"}

# Mapa ticker BVL → ticker Yahoo Finance (para precios en tiempo real)
YAHOO_MAP = {
    "MINSURI1": "MINSURI1.LM", "FERREYC1": "FERREYC1.LM",
    "SIDERC1":  "SIDERC1.LM",  "ORYGENC1": "ORYGENC1.LM",
    "CPACASC1": "CPACASC1.LM", "INRETC1":  "INRETC1.LM",
    "JDOC":     "JDOC.LM",     "JAVA":     "JAVA.LM",
    "VOLCABC1": "VOLCABC1.LM",
}

# ══════════════════════════════════════════════════════════════════════════════
# 1. LECTURA DEL EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def leer_excel(path: Path) -> dict:
    """Lee todas las hojas relevantes del Excel y devuelve datos crudos."""
    print(f"📂 Leyendo {path.name} …")
    wb = openpyxl.load_workbook(path, data_only=True)

    compras   = _leer_hoja(wb, "Compras")
    ventas    = _leer_hoja(wb, "Ventas")
    cerradas  = _leer_hoja(wb, "Cerradas")
    dividendos = _leer_hoja(wb, "Dividendos")
    nav_data  = _leer_hoja(wb, "P&L Total")

    return dict(compras=compras, ventas=ventas,
                cerradas=cerradas, dividendos=dividendos,
                nav_data=nav_data)


def _leer_hoja(wb, nombre: str) -> list[dict]:
    """Lee una hoja como lista de dicts usando la primera fila como headers."""
    if nombre not in wb.sheetnames:
        print(f"  AVISO: hoja '{nombre}' no encontrada")
        return []
    ws = wb[nombre]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col{i}"
               for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: row[i] for i in range(len(headers))})
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 2. CÁLCULO FIFO
# ══════════════════════════════════════════════════════════════════════════════

def calcular_fifo(compras: list, ventas: list) -> tuple[list, list]:
    """
    Aplica FIFO sobre compras y ventas.
    Devuelve (posiciones_abiertas, posiciones_cerradas).

    Campos esperados en compras: ticker, fecha, cantidad, precio, moneda
    Campos esperados en ventas:  ticker, fecha, cantidad, precio, moneda
    (Los nombres exactos de columna se detectan automáticamente)
    """
    # Normalizar nombres de columna
    def campo(row, *opciones):
        for op in opciones:
            for k in row:
                if k and op.lower() in k.lower():
                    return row[k]
        return None

    # Agrupar compras por ticker como colas FIFO
    colas = defaultdict(list)   # ticker → [(qty, precio, moneda), …]
    moneda_map = {}

    for c in compras:
        t   = str(campo(c, "ticker", "symbol", "instrumento") or "").strip().upper()
        qty = campo(c, "cantidad", "quantity", "qty", "acciones")
        px  = campo(c, "precio", "price", "px")
        mon = str(campo(c, "moneda", "currency", "cu") or "USD").strip().upper()
        if not t or qty is None or px is None:
            continue
        qty = float(qty); px = float(px)
        if qty > 0:
            colas[t].append([qty, px, mon])
            moneda_map[t] = mon

    # Procesar ventas FIFO
    realizados = {}   # ticker → {qty, costo_prom, precio_venta, moneda, pnl}

    for v in ventas:
        t   = str(campo(v, "ticker", "symbol", "instrumento") or "").strip().upper()
        qty = campo(v, "cantidad", "quantity", "qty", "acciones")
        px  = campo(v, "precio", "price", "px")
        if not t or qty is None or px is None:
            continue
        qty = float(qty); px = float(px)
        cola = colas.get(t, [])
        restante = qty
        costo_total = 0.0
        while restante > 0 and cola:
            lote_qty, lote_px, lote_mon = cola[0]
            consumido = min(restante, lote_qty)
            costo_total += consumido * lote_px
            restante -= consumido
            cola[0][0] -= consumido
            if cola[0][0] < 1e-6:
                cola.pop(0)
        costo_prom = costo_total / qty if qty else 0
        pnl = (px - costo_prom) * qty
        mon = moneda_map.get(t, "USD")
        if t not in realizados:
            realizados[t] = dict(t=t, q=qty, w=round(costo_prom, 4),
                                 pv=round(px, 4), cu=mon,
                                 pnl=round(pnl, 2),
                                 pnl_pct=round(pnl / (costo_prom * qty), 4) if costo_prom else 0,
                                 d=0.0)
        else:
            realizados[t]["q"] += qty
            realizados[t]["pnl"] += round(pnl, 2)

    cerradas = list(realizados.values())

    # Posiciones abiertas: saldo restante en colas
    abiertas = []
    for t, cola in colas.items():
        if not cola:
            continue
        qty_total = sum(l[0] for l in cola)
        costo_total = sum(l[0] * l[1] for l in cola)
        w = costo_total / qty_total if qty_total else 0
        mon = moneda_map.get(t, "USD")
        abiertas.append(dict(t=t, q=round(qty_total, 0), w=round(w, 4),
                             c=round(costo_total, 2), cu=mon,
                             px=None, pxp=None, val=None, d=0.0))

    return abiertas, cerradas


# ══════════════════════════════════════════════════════════════════════════════
# 3. PRECIOS EN TIEMPO REAL (Yahoo Finance)
# ══════════════════════════════════════════════════════════════════════════════

def obtener_precios(tickers: list[str]) -> tuple[dict, dict]:
    """
    Devuelve (precios_hoy, precios_ayer).
    Para tickers BVL usa el sufijo .LM; para el resto directo.
    Si yfinance no está disponible devuelve dicts vacíos.
    """
    if not YFINANCE_OK:
        return {}, {}

    yahoo_tickers = [YAHOO_MAP.get(t, t) for t in tickers]
    print(f"🌐 Descargando precios ({len(yahoo_tickers)} tickers) …")

    try:
        data = yf.download(yahoo_tickers, period="5d", auto_adjust=True,
                           progress=False, threads=True)
        close = data["Close"]

        precios_hoy  = {}
        precios_ayer = {}
        reverse_map  = {v: k for k, v in YAHOO_MAP.items()}

        for yt in yahoo_tickers:
            col = yt if yt in close.columns else None
            if col is None:
                continue
            serie = close[col].dropna()
            if len(serie) < 1:
                continue
            t_orig = reverse_map.get(yt, yt)
            precios_hoy[t_orig]  = round(float(serie.iloc[-1]), 4)
            precios_ayer[t_orig] = round(float(serie.iloc[-2]), 4) if len(serie) >= 2 else precios_hoy[t_orig]

        print(f"   ✓ {len(precios_hoy)} precios obtenidos")
        return precios_hoy, precios_ayer

    except Exception as e:
        print(f"   AVISO precios: {e}")
        return {}, {}


# ══════════════════════════════════════════════════════════════════════════════
# 4. ENRIQUECER POSICIONES CON PRECIOS Y DIVIDENDOS
# ══════════════════════════════════════════════════════════════════════════════

def enriquecer(abiertas: list, cerradas: list, dividendos: list,
               precios_hoy: dict, precios_ayer: dict,
               px_excel: dict) -> tuple[list, list]:
    """Añade px, pxp, val y dividendos a cada posición."""

    # Acumular dividendos por ticker (en moneda original)
    divs_por_ticker = defaultdict(float)
    for d in dividendos:
        t = str(d.get("Ticker") or d.get("ticker") or "").strip().upper()
        m = d.get("Monto") or d.get("monto") or d.get("Amount") or 0
        if t and m:
            divs_por_ticker[t] += float(m)

    for p in abiertas:
        t = p["t"]
        # Precio: Yahoo > Excel > None
        px  = precios_hoy.get(t)  or px_excel.get(t)
        pxp = precios_ayer.get(t) or px_excel.get(t)
        p["px"]  = round(px, 4)  if px  is not None else None
        p["pxp"] = round(pxp, 4) if pxp is not None else None
        p["val"] = round(p["q"] * px, 2) if px is not None else None
        p["d"]   = round(divs_por_ticker.get(t, 0.0), 2)

    # Dividendos en cerradas
    for p in cerradas:
        p["d"] = round(divs_por_ticker.get(p["t"], 0.0), 2)

    return abiertas, cerradas


# ══════════════════════════════════════════════════════════════════════════════
# 5. CONSTRUIR TODOS LOS BLOQUES DE DATOS PARA EL HTML
# ══════════════════════════════════════════════════════════════════════════════

def construir_datos(raw: dict, abiertas: list, cerradas: list,
                    divs_excel: list, precios_hoy: dict,
                    precios_ayer: dict) -> dict:
    """Construye el dict completo de datos para inyectar en el HTML."""

    # ── DIVS (tabla de dividendos) ──
    def campo(row, *opts):
        for op in opts:
            for k in row:
                if k and op.lower() in k.lower():
                    return row[k]
        return None

    divs = []
    for d in divs_excel:
        t   = str(campo(d, "ticker", "instrumento") or "").strip().upper()
        f   = campo(d, "fecha", "date")
        m   = campo(d, "monto", "amount", "importe")
        cu  = str(campo(d, "moneda", "currency") or "USD").strip().upper()
        if not t or m is None:
            continue
        fecha_str = f.strftime("%Y-%m-%d") if hasattr(f, "strftime") else str(f or "")
        # Fix Minsur: cotiza PEN pero dividend pudo estar en USD-equiv
        if t == "MINSURI1" and cu == "USD":
            m = round(float(m) * PEN_RATE, 2)
            cu = "PEN"
        divs.append({"fecha": fecha_str, "t": t,
                     "m": round(float(m), 2), "cu": cu})
    divs.sort(key=lambda x: x["fecha"], reverse=True)

    # ── KPIS ──
    def val_usd(p):
        if p["cu"] == "PEN":
            return (p["val"] or 0) / PEN_RATE
        return p["val"] or 0

    nav_usd_total = sum(val_usd(p) for p in abiertas if p["cu"] == "USD")
    nav_pen_total = sum(p["val"] or 0 for p in abiertas if p["cu"] == "PEN")
    nav_total     = nav_usd_total + nav_pen_total / PEN_RATE
    cost_total    = sum(p["c"] if p["cu"] == "USD" else p["c"] / PEN_RATE for p in abiertas)

    unrealized = sum(
        ((p["val"] or 0) - p["c"]) if p["cu"] == "USD"
        else ((p["val"] or 0) - p["c"]) / PEN_RATE
        for p in abiertas
    )
    realized_usd = sum(p["pnl"] for p in cerradas if p["cu"] == "USD")
    realized_pen = sum(p["pnl"] for p in cerradas if p["cu"] == "PEN")

    divs_usd = sum(d["m"] for d in divs if d["cu"] == "USD")
    divs_pen = sum(d["m"] for d in divs if d["cu"] == "PEN")
    dividends_total = divs_usd + divs_pen / PEN_RATE

    pnl1d = sum(
        (p["px"] - p["pxp"]) * p["q"] if p["cu"] == "USD"
        else ((p["px"] - p["pxp"]) * p["q"]) / PEN_RATE
        for p in abiertas
        if p["px"] is not None and p["pxp"] is not None
    )

    kpis = {
        "nav_total":      round(nav_total),
        "nav_usd":        round(nav_usd_total),
        "nav_pen":        round(nav_pen_total),
        "cost_total":     round(cost_total),
        "unrealized":     round(unrealized),
        "unrealized_pct": round(unrealized / cost_total, 4) if cost_total else 0,
        "realized":       round(realized_usd + realized_pen / PEN_RATE),
        "realized_usd":   round(realized_usd),
        "realized_pen":   round(realized_pen),
        "dividends":      round(dividends_total),
        "divs_usd":       round(divs_usd),
        "divs_pen":       round(divs_pen),
        "total_pnl":      round(unrealized + realized_usd + realized_pen / PEN_RATE + dividends_total),
        "pnl1d":          round(pnl1d),
    }

    return dict(
        OPEN     = abiertas,
        CLOSED   = cerradas,
        DIVS     = divs,
        KPIS_DATA= kpis,
        PRICES   = precios_hoy,
        PRICES_PREV = precios_ayer,
        updated  = datetime.now().strftime("%d/%m/%Y %H:%M"),
    )


# ══════════════════════════════════════════════════════════════════════════════
# 6. INYECTAR DATOS EN EL HTML TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════

# Marcadores en el template (el template es el HTML original con estos comentarios)
MARKER_START = "/* ══ DATOS_START ══ */"
MARKER_END   = "/* ══ DATOS_END ══ */"

def generar_html(datos: dict) -> str:
    """Lee el template y reemplaza el bloque de datos con los frescos."""
    html = TEMPLATE.read_text(encoding="utf-8")

    # Construir bloque JS
    o  = datos["OPEN"]
    c  = datos["CLOSED"]
    d  = datos["DIVS"]
    k  = datos["KPIS_DATA"]
    px = datos["PRICES"]
    pp = datos["PRICES_PREV"]
    up = datos["updated"]

    def jd(obj): return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))

    open_js = "const OPEN=[\n"
    for p in o:
        open_js += (f'  {{t:{jd(p["t"])},q:{p["q"]},w:{p["w"]},'
                    f'c:{p["c"]},cu:{jd(p["cu"])},px:{p["px"]},'
                    f'pxp:{p["pxp"]},val:{p["val"]},d:{p["d"]}}},\n')
    open_js += "];"

    bloque = f"""
const OPEN={jd(o)};
const CLOSED={jd(c)};
const DIVS={jd(d)};
const KPIS_DATA={jd(k)};
const PRICES={jd(px)};
const PRICES_PREV={jd(pp)};
const LAST_UPDATED={jd(up)};
"""

    # Reemplazar entre marcadores
    if MARKER_START in html and MARKER_END in html:
        start = html.index(MARKER_START)
        end   = html.index(MARKER_END) + len(MARKER_END)
        html  = html[:start] + MARKER_START + bloque + MARKER_END + html[end:]
    else:
        # Fallback: reemplazar el bloque const OPEN…const PRICES_PREV existente
        patron = r"const OPEN=\[[\s\S]*?const PRICES_PREV=\{[^;]*\};"
        reemplazo = bloque.strip()
        html = re.sub(patron, reemplazo, html, count=1)

    return html


# ══════════════════════════════════════════════════════════════════════════════
# 7. SUBIR A GITHUB PAGES
# ══════════════════════════════════════════════════════════════════════════════

def subir_github(mensaje: str = None):
    """Hace git add + commit + push del index.html generado."""
    if mensaje is None:
        mensaje = f"dashboard: actualización {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    cmds = [
        ["git", "-C", str(BASE_DIR), "add", "index.html"],
        ["git", "-C", str(BASE_DIR), "commit", "-m", mensaje],
        ["git", "-C", str(BASE_DIR), "push"],
    ]
    for cmd in cmds:
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode != 0 and "nothing to commit" not in r.stdout:
            print(f"  git: {r.stderr.strip() or r.stdout.strip()}")
        else:
            print(f"  ✓ {' '.join(cmd[2:4])}")


# ══════════════════════════════════════════════════════════════════════════════
# 8. PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def _hash_archivo(path: Path) -> str:
    return hashlib.md5(path.read_bytes()).hexdigest() if path.exists() else ""


def correr(args):
    """Corre el pipeline completo una vez."""
    # 1. Leer Excel
    raw = leer_excel(EXCEL_PATH)

    # 2. FIFO
    abiertas, cerradas_fifo = calcular_fifo(raw["compras"], raw["ventas"])

    # Agregar posiciones cerradas del Excel si existen
    cerradas_excel = []
    for row in raw.get("cerradas", []):
        def c(r, *o):
            for op in o:
                for k in r:
                    if k and op.lower() in k.lower():
                        return r[k]
            return None
        t   = str(c(row, "ticker", "instrumento") or "").strip().upper()
        pnl = c(row, "p&l", "pnl", "ganancia", "resultado")
        pv  = c(row, "precio venta", "venta", "precio_venta")
        w   = c(row, "costo", "precio compra", "costo_prom")
        q   = c(row, "cantidad", "qty")
        cu  = str(c(row, "moneda", "currency") or "USD").upper()
        if t and pnl is not None:
            cerradas_excel.append(dict(
                t=t, q=float(q or 0), w=round(float(w or 0), 4),
                pv=round(float(pv or 0), 4), cu=cu,
                pnl=round(float(pnl), 2),
                pnl_pct=round(float(pnl) / (float(w or 1) * float(q or 1)), 4),
                d=0.0
            ))

    cerradas = cerradas_excel if cerradas_excel else cerradas_fifo

    # 3. Precios
    tickers = [p["t"] for p in abiertas]
    if args.no_precios or not YFINANCE_OK:
        precios_hoy, precios_ayer = {}, {}
    else:
        precios_hoy, precios_ayer = obtener_precios(tickers)

    # Extraer precios del Excel como fallback (de la hoja DATA o Stock)
    px_excel = {}
    for p in abiertas:
        if p.get("px") is not None:
            px_excel[p["t"]] = p["px"]

    # 4. Enriquecer
    abiertas, cerradas = enriquecer(abiertas, cerradas, raw["dividendos"],
                                    precios_hoy, precios_ayer, px_excel)

    # 5. Construir datos
    datos = construir_datos(raw, abiertas, cerradas, raw["dividendos"],
                            precios_hoy or px_excel, precios_ayer or px_excel)

    # 6. Generar HTML
    html = generar_html(datos)
    OUTPUT.write_text(html, encoding="utf-8")
    print(f"✅ HTML generado → {OUTPUT}")

    # 7. Subir a GitHub
    if not args.no_github:
        subir_github()
        print(f"🚀 Publicado en: https://{GITHUB_REPO.split('/')[0]}.github.io/{GITHUB_REPO.split('/')[1]}/")

    return True


def modo_watch(args):
    """Vigila el Excel y corre el pipeline cuando cambia."""
    print(f"👁  Vigilando {EXCEL_PATH.name} …  (Ctrl+C para salir)")
    ultimo_hash = ""
    while True:
        try:
            h = _hash_archivo(EXCEL_PATH)
            if h != ultimo_hash:
                if ultimo_hash:   # no correr en el primer arranque
                    print(f"\n📝 Cambio detectado — {datetime.now().strftime('%H:%M:%S')}")
                ultimo_hash = h
                if ultimo_hash:
                    correr(args)
            time.sleep(10)        # revisar cada 10 segundos
        except KeyboardInterrupt:
            print("\n⏹  Watcher detenido.")
            break
        except Exception as e:
            print(f"  ERROR: {e}")
            time.sleep(30)


# ══════════════════════════════════════════════════════════════════════════════
# PUNTO DE ENTRADA
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Portfolio Dashboard Updater")
    p.add_argument("--watch",       action="store_true", help="Vigilar cambios en el Excel")
    p.add_argument("--no-github",   action="store_true", help="No subir a GitHub")
    p.add_argument("--no-precios",  action="store_true", help="No descargar precios de Yahoo")
    args = p.parse_args()

    if args.watch:
        modo_watch(args)
    else:
        correr(args)
